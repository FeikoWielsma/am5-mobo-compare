import pandas as pd
import warnings
import openpyxl

# Suppress specific warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

EXCEL_FILE = "AM5 Motherboards Sheet (X870_X670_B850_B650_B840_A620).xlsx"
SHEETS_TO_LOAD = ['X870E', 'X670(E)', 'X870', 'B850', 'B650(E)', 'B840', 'A620(A)']

def extract_components(record):
    """
    Extracts structured components from the flat record dictionary.
    Maps exactly to Excel Top-Level Headers, accounting for flattened hierarchy prefixes.
    """
    
    # Helper to safely get value or empty string
    def get(key):
        val = record.get(key, "")
        if val is None: return ""
        return str(val).strip()

    # --- 1. General ---
    general = {
        'brand': get('Brand'),
        'model': get('Model'),
        'chipset': get('Chipset'),
        'form_factor': get('Form Factor'),
        'socket_ren': get('Socket (Ren)'), 
        'dimms': get('DIMMs'), 
    }

    # --- 2. Rear I/O ---
    rear_io = {
        'legacy_ps2': get('Rear I/O|Legacy|PS/2'),
        'usb': {
            'type_a': {
                'v2_0': get('Rear I/O|USB|Type A|2.0'),
                'v3_2_g1': get('Rear I/O|USB|Type A|3.2G1 (5Gbps)'),
                'v3_2_g2': get('Rear I/O|USB|Type A|3.2G2 (10Gbps)'),
                'total': get('Rear I/O|USB|Type A|USBA Total')
            },
            'type_c': {
                'v3_2_g1': get('Rear I/O|USB|Type C|3.2G1 (5Gbps)'),
                'v3_2_g2': get('Rear I/O|USB|Type C|3.2G2 (10Gbps)'),
                'v3_2_g2x2': get('Rear I/O|USB|Type C|3.2G2x2 (20Gbps)'),
                'usb4': get('Rear I/O|USB|Type C|USB4 (40Gbps)'),
                'total': get('Rear I/O|USB|Type C|USB-C Total')
            },
            'total_usb': get('Rear I/O|USB|Total USB')
        },
        'buttons': {
            'flash': get('Rear I/O|Buttons|BIOS Flashback'),
            'clear_cmos': get('Rear I/O|Buttons|Clear CMOS')
        }
    }

    # --- 3. Video Outs ---
    video_outs = {
        'internal': get('Video Outs|Modern|Internal'),
        'usbc_dp': get('Video Outs|Modern|USB-C (DP Altmode)'),
        'hdmi': get('Video Outs|Modern|HDMI'),
        'dp': get('Video Outs|Modern|DP')
    }

    # --- 4. Expansion ---
    expansion = {
        'pcie_slots': {
            'x16': {
                'lanes': get('Expansion|PCIe Slots|Physical x16|Electrical Lanes'),
                'total': get('Expansion|PCIe Slots|Physical x16|Total')
            },
            'x1_x4': {
                'lanes': get('Expansion|PCIe Slots|Physical x1/x4|Electrical Lanes'),
                'total': get('Expansion|PCIe Slots|Physical x1/x4|Total')
            },
            'total_count': get('Expansion|PCIe Slots|Total Slot Count')
        },
        'storage': {
            'sata': get('Expansion|Storage|SATA'),
            'pcie_storage': {
                'm2': get('Expansion|Storage|PCIe Storage|M.2 (M)'),
                'total_m2': get('Expansion|Storage|PCIe Storage|Total M.2'),
                'aic': get('Expansion|Storage|PCIe Storage|AIC'), # Add-in card
                'slimsas': get('Expansion|Storage|PCIe Storage|SlimSAS')
            }
        }
    }

    # --- 5. Memory (FIXED KEYS) ---
    memory = {
        'slots': get('General|Memory|RAM slots'), # Fixed Path
        'capacity': get('General|Memory|Max. capacity'), # Fixed Path
        'max_speed': get('General|Memory|Max speed'), # Guessing, might verify
        'topology': get('General|Memory|Topology') # Guessing
    }

    # --- 6. Network (FIXED KEYS) ---
    network = {
        'lan_controller': get('General|Networking|Ethernet|LAN Controller'), # Likely path if clustered
        'lan_speed': get('General|Networking|Ethernet|LAN'),
        'wifi': get('General|Networking|Wireless Networking|Wi-Fi'),
        'bt': get('General|Networking|Wireless Networking|Bluetooth')
    }
    # Fallback checks if General|Networking doesn't exist but Networking does
    if not network['lan_speed']:
         network['lan_speed'] = get('Networking|Ethernet|LAN')
         network['lan_controller'] = get('Networking|Ethernet|LAN Controller')
         network['wifi'] = get('Networking|Wireless Networking|Wi-Fi')
         network['bt'] = get('Networking|Wireless Networking|Bluetooth')


    # --- 7. Audio (FIXED KEYS) ---
    audio = {
        'codec': get('General|Audio|Audio Codec+DAC'), # Fixed Path
        'features': get('General|Audio|Amp/Feat. etc')
    }

    # --- 8. VRM ---
    vrm = {
        'config': get('VRM & Power|Configuration'), # Header in screenshot says VRM/Power maybe? Or just VRM?
        'mosfets': get('VRM & Power|MOSFETs'),
        'heatsink': get('VRM & Power|Heatsink'),
        'fan': get('VRM & Power|Fan')
    }
    # Fallback to just "VRM" if "VRM & Power" fails
    if not vrm['config']:
        vrm['config'] = get('VRM|Configuration')
        vrm['mosfets'] = get('VRM|MOSFETs')
        vrm['heatsink'] = get('VRM|Heatsink')
        vrm['fan'] = get('VRM|Fan')
    
    # --- 9. Internal I/O ---
    internal_io = {
        'fan_headers': get('Internal I/O|Fan Headers'),
        'rgb_headers': get('Internal I/O|RGB Headers|ARGB/RGB'),
        'usb_front': {
             'c': get('Internal I/O|USB Headers|Type-C (Front)'),
             'v3': get('Internal I/O|USB Headers|USB 3.0'),
             'v2': get('Internal I/O|USB Headers|USB 2.0')
        }
    }

    return {
        'general': general,
        'rear_io': rear_io,
        'video_outs': video_outs,
        'expansion': expansion,
        'memory': memory,
        'network': network,
        'audio': audio,
        'vrm': vrm,
        'internal_io': internal_io
    }

def load_data():
    all_mobos = []
    final_structure = []
    
    try:
        print(f"Loading Excel file: {EXCEL_FILE}...")
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        xl_pd = pd.ExcelFile(EXCEL_FILE, engine='openpyxl')
        
        for sheet_name in SHEETS_TO_LOAD:
            if sheet_name not in wb.sheetnames: 
                continue
                
            print(f"Processing sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # --- HEADER PARSING ---
            # 1. Anchor Row
            data_rows = []
            max_col_to_read = min(ws.max_column, 200)
            for r in range(1, 21):
                row_vals = []
                for c in range(1, max_col_to_read + 1):
                    val = ws.cell(row=r, column=c).value
                    row_vals.append(str(val).strip() if val is not None else "")
                data_rows.append(row_vals)
            
            anchor_row_idx = -1
            for i, row in enumerate(data_rows):
                if "Brand" in row and "Model" in row:
                    anchor_row_idx = i
                    break
            if anchor_row_idx == -1:
                 for i, row in enumerate(data_rows):
                    if "Make" in row and "Board Name" in row:
                        anchor_row_idx = i
                        break
            
            if anchor_row_idx == -1: continue

            primary_header_idx = anchor_row_idx + 1
            
            # Helper to process headers (embedded)
            def process_headers(worksheet, anchors):
                anchor_idx, primary_idx = anchors
                start_row = 0
                max_col = worksheet.max_column
                merged_ranges = list(worksheet.merged_cells.ranges)
                header_matrix = []
                for r_idx in range(start_row, primary_idx + 1):
                    excel_row = r_idx + 1
                    row_data = []
                    for c_idx in range(max_col):
                         excel_col = c_idx + 1
                         cell = worksheet.cell(row=excel_row, column=excel_col)
                         val = cell.value
                         origin = f"C_{excel_row}_{excel_col}"
                         if val is None:
                             for mr in merged_ranges:
                                 if excel_row >= mr.min_row and excel_row <= mr.max_row:
                                     if excel_col >= mr.min_col and excel_col <= mr.max_col:
                                         tl_cell = worksheet.cell(row=mr.min_row, column=mr.min_col)
                                         val = tl_cell.value
                                         origin = f"M_{mr.min_row}_{mr.min_col}"
                                         break
                         clean_val = str(val).strip().replace('\n', ' ') if val is not None else ""
                         row_data.append({'val': clean_val, 'id': origin})
                    header_matrix.append(row_data)

                leaf_row = header_matrix[-1]
                hierarchy_rows = header_matrix[:-1]
                
                valid_hierarchy_rows = []
                for idx, r_data in enumerate(hierarchy_rows):
                    first_val = r_data[0]['val'].lower()
                    if "use the tabs" in first_val or "missing/incorrect" in first_val or "contact me" in first_val: continue
                    if all(x['val'] == "" for x in r_data): continue
                    valid_hierarchy_rows.append(r_data)
                hierarchy_rows = valid_hierarchy_rows
                
                columns_info = [] 
                seen_signatures = set()
                seen_key_counts = {}
                last_valid_col_info = None 
                
                for col_idx in range(len(leaf_row)):
                    col_ids = [row[col_idx]['id'] for row in hierarchy_rows] + [leaf_row[col_idx]['id']]
                    signature = tuple(col_ids)
                    if signature in seen_signatures: continue
                    seen_signatures.add(signature)
                    
                    h_vals = [row[col_idx]['val'] for row in hierarchy_rows]
                    leaf_val = leaf_row[col_idx]['val'].replace('\n', ' ')
                    if leaf_val.lower() == 'nan': leaf_val = ""
                    clean_parents = [p for p in h_vals if p]
                    
                    # Dedupe
                    if clean_parents:
                        dedupped = [clean_parents[0]]
                        for p in clean_parents[1:]:
                            if p != dedupped[-1]: dedupped.append(p)
                        clean_parents = dedupped
                        while clean_parents:
                            last_p = clean_parents[-1].lower()
                            l_name = leaf_val.lower()
                            should_pop = False
                            if l_name.strip() == last_p.strip(): should_pop = True
                            elif len(l_name) > 4 and len(last_p) > 4 and l_name in last_p: should_pop = True
                            if should_pop: clean_parents.pop()
                            else: break

                    is_aux = False
                    if last_valid_col_info and clean_parents == last_valid_col_info['path'] and not leaf_val and last_valid_col_info['name']:
                        is_aux = True
                    if is_aux: continue

                    if not leaf_val:
                        if clean_parents: leaf_val = clean_parents.pop()
                        else: leaf_val = f"Unnamed_{col_idx}"

                    # Overrides (Crucial for consistent keys)
                    if "Rear I/O Image" in leaf_val:
                        clean_parents = ["Rear I/O Image"]
                        leaf_val = "View"
                    elif "Total USB" in leaf_val:
                         # Force sibling status with Type A/C
                        clean_parents = ['Rear I/O', 'USB']
                    elif "Lane-sharing" in leaf_val and "bifurcation" in leaf_val:
                        clean_parents = ["Notes"]
                        leaf_val = "Details"
                    
                    path_str = "|".join(clean_parents + [leaf_val])
                    
                    # Key handling
                    if leaf_val in ["Brand", "Make"]: base_key = "Brand"
                    elif leaf_val in ["Model", "Board Name"]: base_key = "Model"
                    elif leaf_val in ["Chipset", "Form Factor"]: base_key = leaf_val
                    else: base_key = path_str
                    
                    if base_key in seen_key_counts:
                        seen_key_counts[base_key] += 1
                        final_key = f"{base_key}_{seen_key_counts[base_key]}"
                    else:
                        seen_key_counts[base_key] = 1
                        final_key = base_key
                        
                    current_col_info = {'key': final_key, 'name': leaf_val, 'path': clean_parents, 'original_col_idx': col_idx}
                    columns_info.append(current_col_info)
                    last_valid_col_info = current_col_info
                    
                return columns_info

            # Capture Keys
            sheet_cols_info = process_headers(ws, (anchor_row_idx, primary_header_idx))
            
            # --- DATA READING ---
            data_start_idx = primary_header_idx + 1
            df_data = xl_pd.parse(sheet_name, header=None, skiprows=data_start_idx)
            
            valid_indices = [c['original_col_idx'] for c in sheet_cols_info]
            valid_keys = [c['key'] for c in sheet_cols_info]
            
            max_pd_col = len(df_data.columns)
            safe_indices = []
            safe_keys = []
            for idx, k in zip(valid_indices, valid_keys):
                if idx < max_pd_col:
                   safe_indices.append(idx)
                   safe_keys.append(k)
            
            df_data = df_data.iloc[:, safe_indices]
            df_data.columns = safe_keys
            
            if "Model" in df_data.columns:
                df_data = df_data.dropna(subset=["Model"])
                df_data = df_data[df_data["Model"].astype(str).str.strip() != ""]
            if "Brand" in df_data.columns:
                df_data["Brand"] = df_data["Brand"].ffill()
                
            records = df_data.to_dict('records')
            for idx, record in enumerate(records):
                 raw_model = str(record.get('Model', ''))
                 safe_model = raw_model.replace('\n', ' ').strip().replace(' ', '_').replace('/', '-').replace('\\', '-')
                 while '__' in safe_model: safe_model = safe_model.replace('__', '_')
                 record['id'] = f"{sheet_name}_{idx}_{safe_model}"
                 record['Sheet'] = sheet_name
                 clean_record = {k: (v if pd.notna(v) else "") for k, v in record.items()}
                 
                 # === ENRICHMENT POINT ===
                 components = extract_components(clean_record)
                 clean_record.update(components)
                 
                 all_mobos.append(clean_record)

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return [], []
        
    return all_mobos, final_structure

if __name__ == "__main__":
    d, s = load_data()
    print(f"Loaded {len(d)} mobos")
