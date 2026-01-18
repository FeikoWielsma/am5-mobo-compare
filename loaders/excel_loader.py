"""
Main Excel data loading orchestration for AM5 motherboards.

This module coordinates the entire data loading pipeline:
    1. Load Excel workbook
    2. For each sheet: detect headers, parse data, unflatten records
    3. Build canonical header tree structure
    4. Return (motherboards, structure) for database insertion

Functions:
    load_data: Main entry point for loading all motherboard data
"""

import pandas as pd
import openpyxl
import warnings
import re

from .config import EXCEL_FILE, SHEETS_TO_LOAD, SKIP_HEADER_PATTERNS
from .header_parser import (
    find_leaf_header_row,
    determine_header_range,
    parse_multi_level_headers
)
from .data_transformer import (
    unflatten_record,
    build_header_tree,
    clean_record_values,
    calculate_lan_score,
    extract_scorecard
)

# Suppress openpyxl warnings about styles/formatting (we only read data values)
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def load_data():
    """
    Load and parse AM5 motherboard data from Excel file.
    
    Reads the Excel file specified in config, parses complex multi-level headers,
    extracts data rows, and transforms everything into hierarchical JSON structures.
    
    Returns:
        tuple: (motherboards, header_tree)
            motherboards: List of dicts, each containing:
                - id: Unique identifier (sheet_idx_model)
                - brand: Manufacturer name
                - model: Motherboard model name
                - chipset: Chipset name (X870E, B650, etc.)
                - specs: Nested dict with all specifications
            
            header_tree: Nested list/dict structure representing the header hierarchy,
                        used for UI rendering and navigation
                        
    Examples:
        >>> mobos, structure = load_data()
        >>> len(mobos)
        150  # Approximate, depends on Excel file
        >>> mobos[0]['brand']
        'ASUS'
        >>> mobos[0]['specs']['General']['Audio']['Codec']
        'ALC4080'
    """
    all_mobos = []
    final_header_tree = []
    structure_built = False
    
    try:
        print(f"Loading Excel file: {EXCEL_FILE}...")
        # rich_text=True allows reading partially bolded cells as CellRichText objects
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True, rich_text=True)
        
        # Process each sheet
        for sheet_name in SHEETS_TO_LOAD:
            if sheet_name not in wb.sheetnames:
                print(f"  Warning: Sheet '{sheet_name}' not found, skipping")
                continue
            
            print(f"Processing sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Step 1: Find leaf header row (contains "Brand" and "Model")
            leaf_row = find_leaf_header_row(ws)
            if leaf_row == -1:
                print(f"  Warning: Could not find header row in '{sheet_name}', skipping")
                continue
            
            # Step 2: Determine full header range
            start_row, end_row = determine_header_range(ws, leaf_row)
            print(f"  Header block: Rows {start_row}-{end_row} (leaf at {leaf_row})")
            
            # Step 3: Parse headers into column info
            sheet_cols = parse_multi_level_headers(ws, start_row, end_row)
            
            # Filter out junk columns
            valid_cols = [
                col for col in sheet_cols
                if not any(skip.lower() in col['key'].lower() for skip in SKIP_HEADER_PATTERNS)
            ]
            
            # Step 4: Build header tree (once, from first sheet)
            if not structure_built:
                final_header_tree = build_header_tree(valid_cols)
                structure_built = True
            
            # Step 5: Load data rows using openpyxl (to capture comments)
            data_start_row = end_row + 1  # Start reading data after header
            records = []
            
            # Read rows directly from worksheet
            for row_idx in range(data_start_row, ws.max_row + 1):
                record = {}
                has_model = False
                
                # Read each column's value and comment
                for col_info in valid_cols:
                    col_idx = col_info['col_idx'] + 1  # openpyxl uses 1-indexed columns
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    # Get cell value
                    value = cell.value
                    key = col_info['key']
                    
                    # Handle RichText / partially bolded cells
                    if hasattr(value, '__iter__') and not isinstance(value, (str, bytes)):
                        full_str = ""
                        html_str = ""
                        has_any_bold = False
                        for part in value:
                            text_part = ""
                            is_part_bold = False
                            if isinstance(part, str):
                                text_part = part
                            elif hasattr(part, 'text'):
                                text_part = part.text
                                font_part = getattr(part, 'font', None)
                                is_part_bold = font_part.bold if font_part else False
                            
                            full_str += text_part
                            if is_part_bold:
                                # Separate trailing punctuation/whitespace from bold tag
                                match = re.match(r'^(.*?)(\s*,?\s*)$', text_part)
                                if match:
                                    main_text, suffix = match.groups()
                                    if main_text:
                                        html_str += f"<b>{main_text}</b>"
                                    html_str += suffix
                                else:
                                    html_str += f"<b>{text_part}</b>"
                                has_any_bold = True
                            else:
                                html_str += text_part
                                
                        record[key] = full_str
                        record[f"{key}_html"] = html_str
                        if has_any_bold:
                            record[f"{key}_bold"] = True
                    else:
                        record[key] = value
                        if cell.font and cell.font.bold:
                            record[f"{key}_bold"] = True
                    
                    # Check if this is the Model column and has a value
                    if key == 'Model' and record[key] and str(record[key]).strip():
                        has_model = True
                    
                    # Extract comment if present
                    if cell.comment:
                        try:
                            comment_text = cell.comment.text
                            if comment_text:
                                comment_text = comment_text.strip()
                                record[f"{key}_comment"] = comment_text
                        except Exception:
                            pass
                
                # Only add record if it has a Model (skip empty rows)
                if has_model:
                    records.append(record)
            
            # Step 6: Process each motherboard record
            
            for idx, record in enumerate(records):
                # Clean all values (strip, remove newlines, etc.)
                clean_record = clean_record_values(record)
                
                # Extract identity fields
                brand = clean_record.get('Brand', '')
                model = clean_record.get('Model', '')
                chipset = clean_record.get('Chipset', '')
                
                # Extract Form Factor
                form_factor = ""
                for k, v in clean_record.items():
                    if k.lower().endswith('|form factor'):
                        form_factor = v
                        break
                if not form_factor:
                     # Fallback for sheets where it might not be nested or named differently
                     for k, v in clean_record.items():
                         if "form factor" in k.lower():
                             form_factor = v
                             break
                
                # Generate unique ID
                safe_model = model.replace(' ', '_').replace('/', '-').replace('\\', '-')
                unique_id = f"{sheet_name}_{idx}_{safe_model}"
                
                # Unflatten into hierarchical structure
                nested_specs = unflatten_record(clean_record)
                
                # Calculate and inject LAN Score (server-side)
                # Find "LAN Controller" value. Path: Networking -> LAN Controller
                # Since keys vary, we check the flat record first or navigate nested.
                # Flat record keys are like "Networking|LAN Controller"
                
                lan_text = ""
                # Try to find LAN/Ethernet key in flat dict
                # Key is usually "General|Networking|Ethernet|LAN" or contains "LAN"
                for k, v in clean_record.items():
                    if "Networking" in k and ("LAN" in k or "Ethernet" in k):
                        lan_text = v
                        break
                
                # Load lookup (cached)
                lan_lookup = load_lan_lookup()
                
                # NORMALIZE and Store Canonical IDs
                # calculate_lan_score now internally calls normalize, but we want to store the IDs too.
                from .data_transformer import normalize_lan_controller
                canonical_controllers = normalize_lan_controller(lan_text, list(lan_lookup.keys()))
                
                # Score is sum of speeds of these controllers
                lan_score = sum(lan_lookup.get(c, 0) for c in canonical_controllers)
                
                # Inject into nested specs
                nested_specs['_lan_score'] = lan_score
                nested_specs['_lan_ids'] = canonical_controllers # Store logical IDs for DB/UI
                
                # Extract Scorecard Data
                scorecard = extract_scorecard(clean_record)
                
                # Inject LAN Badges (Consistency with Frontend)
                from .data_transformer import inject_scorecard_lan_badges
                inject_scorecard_lan_badges(scorecard, canonical_controllers, lan_lookup)
                
                nested_specs['_scorecard'] = scorecard
                
                # Create motherboard record
                mobo_record = {
                    'id': unique_id,
                    'brand': brand,
                    'model': model,
                    'chipset': chipset,
                    'form_factor': form_factor,
                    'specs': nested_specs
                }
                
                all_mobos.append(mobo_record)
            
            print(f"  Loaded {len(records)} motherboards from '{sheet_name}'")
    
    except Exception as e:
        print(f"Error loading data: {e}")
        import traceback
        traceback.print_exc()
        return [], []
    
    print(f"\nTotal motherboards loaded: {len(all_mobos)}")
    return all_mobos, final_header_tree


from functools import lru_cache

@lru_cache(maxsize=1)
def load_lan_lookup():
    """
    Load LAN Controller speed mapping from the 'About' sheet.
    Range F8:G20.
    Returns: dict { 'normalized_name': speed_in_mbps }
    """
    print("Loading LAN lookup (uncached)...")
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        if "About" not in wb.sheetnames:
            print("Warning: 'About' sheet for LAN lookup not found.")
            return {}
        
        ws = wb["About"]
        lookup = {}
        
        # Rows 8 to 20 approx, but let's go until empty
        # F is col 6, G is col 7
        for row in range(8, 25): # Safety buffer
            name_cell = ws.cell(row=row, column=6)
            speed_cell = ws.cell(row=row, column=7)
            
            name = name_cell.value
            speed_str = speed_cell.value
            
            if not name:
                continue
                
            name = str(name).strip()
            if not speed_str:
                continue
                
            speed_str = str(speed_str).upper()
            
            # Parse speed
            # "Double 25G" -> 50000
            # "2.5G" -> 2500
            # "1G" -> 1000
            
            multiplier = 1
            if "DOUBLE" in speed_str or "DUAL" in speed_str:
                multiplier = 2
            
            base_speed = 0
            if "25G" in speed_str:
                base_speed = 25000
            elif "10G" in speed_str:
                base_speed = 10000
            elif "2.5G" in speed_str:
                base_speed = 2500
            elif "5G" in speed_str:
                base_speed = 5000
            elif "1G" in speed_str:
                base_speed = 1000
            
            total_speed = base_speed * multiplier
            
            if total_speed > 0:
                lookup[name] = total_speed
                
        return lookup
        
    except Exception as e:
        print(f"Error loading LAN lookup: {e}")
        return {}


if __name__ == "__main__":
    # Test loading
    mobos, structure = load_data()
    print(f"\nLoaded {len(mobos)} motherboards")
    
    if mobos:
        print("\nSample motherboard:")
        sample = mobos[0]
        print(f"  ID: {sample['id']}")
        print(f"  Brand: {sample['brand']}")
        print(f"  Model: {sample['model']}")
        print(f"  Chipset: {sample['chipset']}")
        print(f"  Specs keys: {list(sample['specs'].keys())[:5]}...")
    
    if structure:
        import json
        print("\nHeader tree sample (first 2 nodes):")
        print(json.dumps(structure[:2], indent=2))
