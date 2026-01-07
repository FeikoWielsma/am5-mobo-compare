import pytest
import os
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import sys
from unittest.mock import patch

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from loaders.excel_loader import load_data
from loaders.data_transformer import unflatten_record
from models.database import DotWrapper

# Mock config to use a temp file
@pytest.fixture
def temp_excel_file(tmp_path):
    """Create a temporary Excel file with known structure."""
    file_path = tmp_path / "test_mobos.xlsx"
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active
    ws.title = "AM5 Motherboards"
    
    # 1. Setup Headers
    # Row 1-2: Top level headers (merged)
    # Row 3: Leaf headers
    
    # Simple structure:
    # Brand | Model | Chipset | General -> Audio -> Codec | Power -> VRM
    
    # Row 1
    ws.cell(row=1, column=1, value="Identity")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=3) # Brand, Model, Chipset
    
    ws.cell(row=1, column=4, value="General")
    ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=4) # Audio
    
    ws.cell(row=1, column=5, value="Power")
    
    # Row 2
    ws.cell(row=2, column=4, value="Audio")
    ws.cell(row=2, column=5, value="VRM configuration")
    
    # Row 3 (Leaf)
    headers = ["Brand", "Model", "Chipset", "Codec", "VRM (VCore)"]
    for idx, h in enumerate(headers, 1):
        ws.cell(row=3, column=idx, value=h)
        
    # 2. Add Data
    data = [
        ("ASUS", "Test Board 1", "X670E", "ALC4080", "16+2"),
        ("MSI", "Test Board 2", "B650", "ALC1200", "12+2+1")
    ]
    
    for row_idx, row_data in enumerate(data, 4):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Add a comment to the VRM cell of the first row
            if row_idx == 4 and col_idx == 5:
                cell.comment = openpyxl.comments.Comment("Good VRM", "Reviewer")

    wb.save(file_path)
    return str(file_path)

def test_full_pipeline_flow(temp_excel_file):
    """
    Test the entire pipeline:
    1. Load Excel
    2. Parse Headers
    3. Unflatten Data
    4. Verify nesting and comments
    """
    
    # Patch EXCEL_FILE to point to our temp file
    # And patch SHEETS_TO_LOAD to match our temp file
    with patch('loaders.excel_loader.SHEETS_TO_LOAD', ['AM5 Motherboards']), \
         patch('loaders.excel_loader.EXCEL_FILE', temp_excel_file):
         
         mobos, structure = load_data()
         
    assert len(mobos) == 2
    
    # Verify First Mobo
    m1 = mobos[0]
    assert m1['brand'] == "ASUS"
    assert m1['model'] == "Test Board 1"
    assert m1['chipset'] == "X670E"
    
    # Verify specs nesting
    specs = m1['specs']
    assert 'General' in specs
    assert 'Audio' in specs['General']
    assert specs['General']['Audio']['Codec'] == "ALC4080"
    
    assert 'Power' in specs
    assert 'VRM configuration' in specs['Power']
    assert specs['Power']['VRM configuration']['VRM (VCore)'] == "16+2"
    
    # Verify Comments
    # The loader adds _comment suffix keys
    vrm_key = "Power|VRM configuration|VRM (VCore)" # Flat key? No, unflattened.
    # Wait, unflatten_record handles the nesting.
    # Let's check how comments are stored. 
    # excel_loader.py lines 134: record[f"{key}_comment"] = comment_text
    # key is the flattened key string (e.g. "General|Audio|Codec" or just header name from col info?)
    # header_parser constructs the keys.
    # We should verify the comment exists in the nested structure
    
    # Verify DotWrapper Access (Simulation of Template Usage)
    dot = DotWrapper(specs)
    
    # Fuzzy access
    assert dot.general.audio.codec == "ALC4080"
    assert dot.power.vrm_configuration.vrm_vcore == "16+2"
    
    # Check comment access?
    # Comments are usually stored with a suffix in the same dict
    # If the key was "VRM (VCore)", the comment key is "VRM (VCore)_comment"
    # Unflattening might nest it similarly?
    # If "Power|VRM|Val" is 10, "Power|VRM|Val_comment" is "Note".
    # So `dot.power.vrm.val_comment` should exist?
    
    # Let's inspect the specs keys to be sure about comment placement
    vrm_node = specs['Power']['VRM configuration']
    # If unflatten logic groups by pipe, and comment key was "Power|VRM...VCore_comment"
    # It should be sibling to the value.
    assert 'VRM (VCore)_comment' in vrm_node
    assert vrm_node['VRM (VCore)_comment'] == "Good VRM"
    
    # Test DotWrapper access to comment
    assert dot.power.vrm_configuration.vrm_vcore_comment == "Good VRM"
